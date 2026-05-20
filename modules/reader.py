from pathlib import Path
import pandas as pd
from modules.utils import safe_str

# Mapeamento dos nomes reais da planilha para os nomes internos do script
MAPA_COLUNAS = {
    "Menu ":       "Menu",        # coluna com espaço extra
    "Menu":        "Menu",
    "Observações": "Descrição",   # coluna de descrição na planilha real
    "Descrição":   "Descrição",
    "Tipo":        "Tipo",
    "Específico":  "Especifico",
    "Especifico":  "Especifico",
    "Liberado":    "Liberado",
}

# Valores considerados como "True" na coluna Liberado
LIBERADO_TRUE = {"true", "verdadeiro", "1", "sim", "yes"}

COLUNAS_OBRIGATORIAS = ["Tipo", "Menu", "Descrição"]


def _detectar_aba(wb_path: Path):
    """
    Retorna o nome da aba que contém a coluna 'Tipo'.
    Prioriza a aba mais recente (última que tiver a coluna).
    """
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    aba_encontrada = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        headers_norm = [str(h).strip() if h else "" for h in headers]
        if "Tipo" in headers_norm:
            aba_encontrada = sheet_name
    wb.close()
    return aba_encontrada


def ler_planilha(caminho: Path):
    """
    Lê .xlsm ou .xlsx, detecta automaticamente a aba correta,
    mapeia os nomes de colunas e retorna apenas linhas válidas.
    """
    sufixo = Path(caminho).suffix.lower()
    avisos = []

    if sufixo == ".csv":
        df = pd.read_csv(caminho, dtype=str)
    elif sufixo not in (".xlsm", ".xlsx", ".xls"):
        raise ValueError(f"Formato não suportado: {sufixo}")
    else:
        # Detecta a aba com coluna Tipo
        aba = _detectar_aba(caminho)
        if not aba:
            raise ValueError(
                "Nenhuma aba com coluna 'Tipo' encontrada na planilha."
            )
        avisos.append(f"Aba utilizada: '{aba}'")
        df = pd.read_excel(caminho, sheet_name=aba, dtype=str, engine="openpyxl")

    # Normaliza e mapeia nomes de colunas
    df.columns = [c.strip() for c in df.columns]
    df = df.rename(columns=MAPA_COLUNAS)
    df = df.fillna("")

    # Checa colunas obrigatórias após mapeamento
    ausentes = [c for c in COLUNAS_OBRIGATORIAS if c not in df.columns]
    if ausentes:
        raise ValueError(
            f"Colunas obrigatórias não encontradas: {', '.join(ausentes)}\n"
            f"Colunas presentes: {', '.join(df.columns)}"
        )

    # Remove linhas completamente vazias
    antes = len(df)
    df = df[
        df["Tipo"].str.strip().ne("") |
        df["Menu"].str.strip().ne("") |
        df["Descrição"].str.strip().ne("")
    ].copy()

    # Filtra apenas Melhoria e Correção pela coluna Tipo.
    # As colunas "FAQ" e "Melhoria" são ignoradas intencionalmente —
    # melhorias entram nas FAQs específicas escritas separadamente.
    tipos_validos = df["Tipo"].str.strip().str.lower().isin(["melhoria", "correção"])
    ignorados = (~tipos_validos & df["Tipo"].str.strip().ne(""))
    if ignorados.sum():
        tipos_ignorados = df.loc[ignorados, "Tipo"].unique().tolist()
        avisos.append(f"Tipos ignorados (não são Melhoria/Correção): {tipos_ignorados}")

    df = df[tipos_validos].reset_index(drop=True)

    # Filtra linhas onde Liberado = True (não devem entrar na FAQ)
    if "Liberado" in df.columns:
        liberado_true = df["Liberado"].str.strip().str.lower().isin(LIBERADO_TRUE)
        qtd_liberados = liberado_true.sum()
        df = df[~liberado_true].reset_index(drop=True)
        if qtd_liberados:
            avisos.append(f"{qtd_liberados} linha(s) ignorada(s) por estarem marcadas como Liberado=True")

    if antes - len(df) > 0:
        avisos.append(f"{antes - len(df)} linhas ignoradas (vazias, tipo inválido ou liberadas)")

    return df, avisos
