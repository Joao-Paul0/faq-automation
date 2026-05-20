"""
Gera a nova planilha copiando o arquivo original via XML (zipfile),
preservando fórmulas, inlineStr, estilos e validações exatamente
como estão. Limpa apenas os dados das linhas (mantém o header intacto).
Converte colunas com TRUE()/FALSE() para checkboxes nativos do Excel 365.
"""
from pathlib import Path
import zipfile
import shutil
import re


def _detectar_sheet_info(zip_path: Path) -> tuple:
    """
    Retorna (sheet_number, sheet_name) da aba que contém coluna 'Tipo'.
    Usa openpyxl para leitura confiável dos headers, depois mapeia
    o nome da aba para o número do sheet via XML.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(zip_path), read_only=True, data_only=True)
    aba_encontrada = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        if "Tipo" in headers:
            aba_encontrada = sheet_name
    wb.close()

    if not aba_encontrada:
        raise ValueError("Nenhuma aba com coluna 'Tipo' encontrada na planilha.")

    # Mapear nome da aba → número do sheet via XML
    with zipfile.ZipFile(str(zip_path)) as z:
        wb_xml   = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        rels_raw = re.findall(
            r'Id="([^"]+)"[^>]*Target="[^"]*?sheet(\d+)\.xml"|Target="[^"]*?sheet(\d+)\.xml"[^>]*Id="([^"]+)"',
            rels_xml
        )
        rels = {}
        for m in rels_raw:
            rid, num = (m[0], m[1]) if m[0] else (m[3], m[2])
            rels[rid] = num

        sheets = re.findall(r'name="([^"]+)"[^>]+r:id="([^"]+)"', wb_xml)
        for name, rid in sheets:
            if name == aba_encontrada:
                num = rels.get(rid)
                if num:
                    return int(num), name

    raise ValueError(f"Aba '{aba_encontrada}' encontrada mas não mapeada no XML.")


# Nomes de colunas que devem virar checkbox
COLUNAS_CHECKBOX = {"Liberado", "FAQ Versao", "Específico"}

def _detectar_colunas_booleanas(sheet_xml: str) -> dict:
    """
    Detecta colunas pelo nome do header (Liberado, FAQ Versao)
    que devem virar checkboxes. Retorna { 'G': True, 'J': True, ... }
    """
    import html
    header_row = re.search(r'<row r="1".*?</row>', sheet_xml, re.DOTALL)
    if not header_row:
        return {}
    
    colunas = {}
    cells = re.findall(r'<c r="([A-Z]+)1"[^>]*>.*?<(?:is><t[^>]*>|v>)([^<]+)', header_row.group(), re.DOTALL)
    for col, val in cells:
        nome = html.unescape(val).strip()
        if nome in COLUNAS_CHECKBOX:
            colunas[col] = True
    return colunas


def _coletar_valores_booleanos(sheet_xml: str, col: str) -> dict:
    """
    Coleta valores TRUE/FALSE de uma coluna por número de linha.
    """
    valores = {}
    for row_match in re.finditer(r'<row r="(\d+)".*?</row>', sheet_xml, re.DOTALL):
        row_num = int(row_match.group(1))
        if row_num == 1:
            continue
        cell = re.search(rf'<c r="{col}{row_num}".*?</c>', row_match.group(0), re.DOTALL)
        if cell:
            valores[row_num] = 'TRUE' if 'TRUE()' in cell.group() else 'FALSE'
    return valores


def _converter_celula_para_checkbox(xml: str, col: str, row_num: int, valor: str) -> str:
    """
    Substitui a célula fórmula por célula booleana pura (t="b").
    """
    v = '1' if valor == 'TRUE' else '0'
    novo = f'<c r="{col}{row_num}" t="b"><v>{v}</v></c>'
    old = re.search(rf'<c r="{col}{row_num}".*?</c>', xml, re.DOTALL)
    if old:
        xml = xml[:old.start()] + novo + xml[old.end():]
    return xml


def _adicionar_extlst_checkboxes(sheet_xml: str, colunas_max: dict) -> str:
    """
    Adiciona o extLst com checkboxControls para cada coluna booleana.
    Compatível com Excel 365 / Excel Online.
    """
    if not colunas_max:
        return sheet_xml

    controles = ""
    for col, max_row in sorted(colunas_max.items()):
        controles += f'\n      <xcb:checkboxControl linkedCell="" defaultValue="0" sqref="{col}2:{col}{max_row}"/>'

    checkbox_ext = f'''<extLst>
  <ext xmlns:xcb="http://schemas.microsoft.com/office/spreadsheetml/2022/formulas" uri="{{7E03D99C-DC04-49d9-9315-930204A7B6E9}}">
    <xcb:checkboxControls>{controles}
    </xcb:checkboxControls>
  </ext>
</extLst>'''

    return sheet_xml.replace('</worksheet>', checkbox_ext + '\n</worksheet>')


def _limpar_dados_sheet(sheet_xml: str) -> str:
    """
    Remove todas as linhas de dados (row r >= 2), mantendo o header (row r="1").
    Converte células booleanas para checkbox com valor 0 (desmarcado).
    """
    sheet_xml = re.sub(r'<row r="[2-9]".*?</row>\s*', '', sheet_xml, flags=re.DOTALL)
    sheet_xml = re.sub(r'<row r="\d{2,}".*?</row>\s*', '', sheet_xml, flags=re.DOTALL)
    return sheet_xml


def gerar_planilha_nova_versao(caminho_origem: Path, nova_versao: str, caminho_saida: str):
    caminho_origem = Path(caminho_origem)
    caminho_saida  = Path(caminho_saida)

    # Detectar qual sheet tem coluna Tipo
    sheet_num, sheet_name_orig = _detectar_sheet_info(caminho_origem)

    # Copiar o arquivo original inteiro
    shutil.copy2(str(caminho_origem), str(caminho_saida))

    # Ler todos os arquivos do zip
    with zipfile.ZipFile(str(caminho_saida), 'r') as z_in:
        arquivos = {name: z_in.read(name) for name in z_in.namelist()}

    sheet_key = f"xl/worksheets/sheet{sheet_num}.xml"
    sheet_xml = arquivos[sheet_key].decode("utf-8")

    # 1. Detectar colunas com TRUE()/FALSE() antes de limpar
    colunas_bool = _detectar_colunas_booleanas(sheet_xml)

    # 2. Limpar dados (remover linhas >= 2)
    sheet_xml = _limpar_dados_sheet(sheet_xml)

    # 3. Adicionar extLst com checkboxes para as colunas detectadas
    if colunas_bool:
        # max_row 1000 para cobrir futuras linhas da nova planilha
        colunas_max = {col: 1000 for col in colunas_bool.keys()}
        sheet_xml = _adicionar_extlst_checkboxes(sheet_xml, colunas_max)

    arquivos[sheet_key] = sheet_xml.encode("utf-8")

    # 4. Atualizar nome da aba no workbook.xml
    wb_xml = arquivos["xl/workbook.xml"].decode("utf-8")
    wb_xml = wb_xml.replace(
        f'name="{sheet_name_orig}"',
        f'name="{nova_versao}"'
    )
    arquivos["xl/workbook.xml"] = wb_xml.encode("utf-8")

    # Reescrever o zip
    with zipfile.ZipFile(str(caminho_saida), 'w', zipfile.ZIP_DEFLATED) as z_out:
        for name, data in arquivos.items():
            z_out.writestr(name, data)
